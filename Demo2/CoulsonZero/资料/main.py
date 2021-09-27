import openpyxl as xl
from openpyxl.chart import BarChart, Reference  #导入条形图和参考


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet =wb["Sheet1"]
    #cell = sheet['C1']
    #print(cell.value)
    #cell = sheet.cell(1,3)
    #print(cell.value)    #同5
    #print(sheet.max_row)

    for row in range(1, sheet.max_row + 1):  #输出C3值
        cell = sheet.cell(row, 4)
        corrected_value = cell.value * 0.9
        corrected_value_cell = sheet.cell(row, 6)
        corrected_value_cell.value = corrected_value
        #print(cell.value)

    values = Reference(sheet,
                       min_row=1,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()   #图表类型
    chart.add_data(values)    #添加图标数据
    sheet.add_chart(chart,'d3')   #添加图标
    wb.save(filename)
